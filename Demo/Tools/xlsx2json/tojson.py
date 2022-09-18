# coding=utf-8

#   作者: mmc
#   日期: 2020/3/19

#   说明:
#       这个py的目的是实现一个在手游开发行业相对通用的Excel=>Json转换工具.
#       支持以下类型:
#           list  []
#           dict  {}
#           type  <>
#           bool
#           number
#           string
#       所有类型都可嵌套组合新的数据类型
#
#   用例:
#   基本类型
# bool b      number n        string s
# 1           1               "a"

#   数组类型
# [bool] b_list   [number] n_list     [string] s_list
# [1, 1, 0]       [1, 2, 3]           ["1", "2", "3"]

#   字典类型
# {bool} b_dict               {number} n_dict             {string} s_dict
# {"1": 1, "2": 0, "3": 1}    {"1": 1, "2": 2, "3": 3}    {"1": "1", "2": "2", "3": "3"}

#   自定义类型
# <bool b, number n, string s> t
# <1, 1, "a">

import sys
import math
import time
import openpyxl

def is_number(str):
    try:
        float(str)
    except ValueError:
        return False
    return True

#   读取单元格文本
def XLGetValue(xlsx, cur_row, cur_col):
    return str(xlsx.cell(cur_row, cur_col).value)

#   跳过空字符
def Skip(buffer, i, l):
    while i != l:
        if ord(buffer[i]) > 32:
            break
        i = i + 1
    return i

#   跳过注释行
def SkipLine(xlsx, cur_row):
    while cur_row != xlsx.max_row:
        if XLGetValue(xlsx, cur_row, 1) != "//":
            break
        cur_row = cur_row + 1
    return cur_row

def SplitTypeName(value):
    split = value.split()
    return len(split) == 1 \
        and (split[0], None) \
        or (split[0], split[1])

#   字符串比较
def EqualString(str0, str1, i):
    l0 = len(str0)
    l1 = len(str1)
    if l0 >= l1:
        return str0[i:i+l1] == str1
    return False

#   解析器
class ParseUnit:
    def __init__(self, func = None, name = None):
        self.mFunc = func
        self.mName = name
        self.mChildren = []
    
    def Append(self, unit):
        self.mChildren.append(unit)

    def SetName(self, name):
        self.mName = name

    def GetName(self):
        return self.mName
    
    def GetType(self):
        if self.mFunc == OnParserList:
            return "list"
        elif self.mFunc == OnParserDict:
            return "dict"
        elif self.mFunc == OnParserType:
            return "type"
        elif self.mFunc == OnParserBool:
            return "bool"
        elif self.mFunc == OnParserInt:
            return "int"
        elif self.mFunc == OnParserNumber:
            return "number"
        elif self.mFunc == OnParserString:
            return "string"

    def GetChildren(self):
        return self.mChildren

    def Exec(self, value, i, l):
        return self.mFunc(value, i, l, self.mChildren)

def ValToKey(value):
    return value[0] != "\"" and "\"" + value + "\"" or value

def OnParserBool(value, i, l, unit_child):
    assert is_number(value[i]),"[bool]值错误 %s" % value[i]
    return i + 1, value[i] == "0" and "false" or "true"

def OnParserInt(value, i, l, unit_child):
    num = []
    dot = False
    while i != l:
        if "0" <= value[i] and "9" >= value[i]:
            num.append(value[i]); i = i + 1
        elif len(num) == 0 and "-" == value[i]:
            num.append(value[i]); i = i + 1
        else:
            break
    return i, "".join(num)

def OnParserNumber(value, i, l, unit_child):
    num = []
    dot = False
    while i != l:
        if "0" <= value[i] and "9" >= value[i]:
            num.append(value[i]); i = i + 1
        elif len(num) == 0 and "-" == value[i]:
            num.append(value[i]); i = i + 1
        elif not dot and "." == value[i]:
            num.append(value[i]); i = i + 1
        else:
            break
    return i, "".join(num)

def OnParserString(value, i, l, unit_child):
    assert value[i] == "\"", "[string]值错误: 开头缺少\""
    i = i + 1
    str = [ ]
    mis = False
    while i != l:
        if value[i] == "\\": mis = True
        if value[i] == "\"" and not mis:
            break
        str.append(value[i])
        i   = i + 1
        mis = False
    assert value[i] == "\"", "[string]值错误: 结尾缺少\""
    return i + 1, "\"" + "".join(str) + "\""

def OnParserList(value, i, l, unit_child):
    assert value[i] == "[", "[list]值错误: 开头缺少["
    ele = [ ]
    while i != l:
        if value[i] == "]": break
        i       = Skip(value, i + 1, l)
        if value[i] == "]": break
        i, data = unit_child[0].Exec(value, i, l)
        i       = Skip(value, i, l)
        ele.append(data)
        assert value[i] == "," or value[i] == "]", "[list]值错误: 缺少,"
    assert value[i] == "]", "[list]值错误: 结尾缺少]"
    return i + 1, "[" + ", ".join(ele) + "]"

def OnParserDict(value, i, l, unit_child):
    assert value[i] == "{", "[dict]值错误: 开头缺少{"
    ele = [ ]
    while i != l:
        if value[i] == "}": break
        i = Skip(value, i + 1, l)
        if value[i] == "}": break
        #   解析Key
        i, key = unit_child[0].Exec(value, i, l)
        i      = Skip(value, i, l)
        assert value[i] == ":", "[dict]值错误: 缺少:"
        i      = Skip(value, i + 1, l)
        i, val = unit_child[1].Exec(value, i, l)
        ele.append("%s: %s" %(key, val))
        assert value[i] == "," or value[i] == "}", "[dict]值错误: 缺少,"
    assert value[i] == "}", "[dict]值错误: 结尾缺少}"
    return i + 1, "{" + ", ".join(ele) + "}"

def OnParserType(value, i, l, unit_child):
    assert value[i] == "<", "[type]值错误: 开头缺少<"
    ele = [ ]
    idx = 0
    while i != l:
        if value[i] == ">": break
        i = Skip(value, i + 1, l)
        #   解析Key
        key     = unit_child[idx].GetName()
        i, val  = unit_child[idx].Exec(value, i, l)
        i       = Skip(value, i, l) ; idx = idx + 1
        ele.append("\"%s\": %s" % (key, val))
        assert value[i] == "," or value[i] == ">", "[type]值错误: 缺少,"
    assert value[i] == ">", "[type]值错误: 结尾缺少>"
    return i + 1, "{" + ", ".join(ele) + "}"

def ParseTypeName(value, i, l):
    b = i
    while i != l and (\
        ("a" <= value[i] and "z" >= value[i]) or \
        ("A" <= value[i] and "Z" >= value[i]) or \
        ("0" <= value[i] and "9" >= value[i]) or \
        ("_" == value[i])):
        i = i + 1
    return i, value[b: i]

def BuildFromValue(value, i, l):
    unit = None
    if EqualString(value, "bool", i):
        i = i + len("bool")
        unit = ParseUnit(OnParserBool)
    elif EqualString(value, "int", i):
        i = i + len("int")
        unit = ParseUnit(OnParserInt)
    elif EqualString(value, "number", i):
        i = i + len("number")
        unit = ParseUnit(OnParserNumber)
    elif EqualString(value, "string", i):
        i = i + len("string")
        unit = ParseUnit(OnParserString)
    elif value[i] == "[":
        i, child = BuildFromValue(value, Skip(value, i + 1, l), l)
        i        = Skip(value, i, l)
        unit = ParseUnit(OnParserList)
        unit.Append(child)
        assert value[i] == "]", "类型解析错误: 数组缺少]"
        i = i + 1
    elif value[i] == "{":
        unit_key = ParseUnit(OnParserString)
        i, child = BuildFromValue(value, Skip(value, i + 1, l), l)
        i        = Skip(value, i, l)
        unit = ParseUnit(OnParserDict)
        unit.Append(unit_key)
        unit.Append(child)
        assert value[i] == "}", "类型解析错误: 字典缺少}"
        i = i + 1
    elif value[i] == "<":
        unit = ParseUnit(OnParserType)
        while i != l:
            if value[i] == ">": break
            i = Skip(value, i + 1, l)
            i, child = BuildFromValue(value, i, l)
            i = Skip(value,i,l);unit.Append(child)
            assert value[i] == "," or value[i] == ">", "类型解析错误: 类型缺少,"
        assert value[i] == ">", "类型解析错误: 类型缺少>"
        i = i + 1

    i       = Skip(value, i, l)
    i, name = ParseTypeName(value, i, l)
    unit.SetName(name)
    return i, unit

#   编译解析器
def Build(xlsx, max_col, cur_row, parse_funcs):
    for cur_col in range(1, max_col + 1):
        try:
            cur_val = XLGetValue(xlsx, cur_row, cur_col)
            x,y = BuildFromValue(cur_val,0,len(cur_val))
            parse_funcs.append(y)
        except AssertionError as e:
            assert False, "%d:%d | %s" % (cur_row, cur_col, e)
    return cur_row + 1

#   解析单元格
def Parse(xlsx, max_col, cur_row, parse_funcs, out):
    for cur_row in range(cur_row, xlsx.max_row + 1):
        if XLGetValue(xlsx, cur_row, 1) == "//":
            continue

        lines = []
        for cur_col in range(1, max_col + 1):
            try:
                value = XLGetValue(xlsx, cur_row, cur_col)
                l = len(value)
                k = parse_funcs[cur_col - 1].GetName()
                _,v=parse_funcs[cur_col - 1].Exec(value, 0, l)
                lines.append("\"%s\": %s" % (k, v))
            except AssertionError as e:
                assert False, "%d:%d | %s" % (cur_row, cur_col, e)
        key = XLGetValue(xlsx, cur_row, 1)
        val = "{" + ", ".join(lines) + "}"
        out.append("%s: %s" % (ValToKey(key), val))
    return cur_row

#   导出Json
def ToJson(ifile):
    clock = time.time()
    #   读第一张表
    xlsx = openpyxl.load_workbook(
            ifile,data_only = True)
    xlsx = xlsx[xlsx.sheetnames[0]]
    #   跳过注释行
    parse_funcs = []
    json_output = []
    max_col = xlsx.max_column
    cur_row = SkipLine(xlsx, 1)

    try:
        print("> Export %s" % ifile)
        cur_row = Build(xlsx, max_col, cur_row, parse_funcs)
        cur_row = Parse(xlsx, max_col, cur_row, parse_funcs, json_output)
        # print("> %.3fs From %s" % (time.time() - clock, ifile))
        return "{\n" + ",\n".join(json_output) + "\n}", parse_funcs
    except AssertionError as e:
        assert False, "%s | %s" % (ifile, e)

def ExportToFile(ifile, ofile):
    with open(ofile, "w") as f:
        f.write((ToJson(ifile))[0])

if __name__ == "__main__":
    ExportToFile(sys.argv[1], sys.argv[2])
